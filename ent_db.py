import re
import openpyxl
from openpyxl import Workbook, load_workbook

#o usuario faz a escolha entre cliente ou empresa
while True:
    print("Você quer cadastrar uma empresa ou deseja sem um cliente?")
    print("Digite 1-Para cadastrar uma empresa")
    print("Digite 2-Para se cadastrar como cliente")
    
    esclog = input("Digite o número da opção desejada: ")
    if esclog == "1":
        nomesoc = input(str("digite o nome da sua empresa: "))
        
        cnpj = input(str("digite o cnpj da sua enpresa: "))
        
        local = input("digite o endereço da sua empresa: ")
        
        num_endere_empre = input("digite o numero do endereço: ")
        
        lis_comp_empre = ['sim', 'não']
        while True:
            print("O enderço da sua empresa tem complemento?")
            for i, comp_empre in enumerate(lis_comp_empre, start=1):
                print(f"{i}. {comp_empre}")
            esc_comp = int(input("digite 1 se tiver complemento, se não 2"))
            
            if 1 <= esc_comp <= len(lis_comp_empre):
                comple_sele = lis_comp_empre[esc_comp-1]
                if comple_sele == 'sim':
                    comp_empre = input("digite o complemento: ")
                    print(f"Endereço: {local}, {num_endere_empre}, complemento: {comp_empre}")
                    
                else:
                    print(f"Endereço: {local}, {num_endere_empre}")
                break
            else:
                print("Escolha inválida. Por favor, selecione 1 ou 2")
                
        while True:
        #solicita um E-mail
            email_empre = input("digite o email da sua empresa: ")
            valid_email_empre = re.compile(r'^[^\s]+@gmail.com$')
            if valid_email_empre.match(email_empre):
                print("email valido")
                break
            else:
                print("Formato do email invalido, Certifique-se de que não há espaços e que termina com @gmail.com.")
                
        numero_tele_empre = input("Digite seu numero com ddd: ")
        
        print("Cadastro comcluido suas informações são:")
        print('''Nome da empresa: {}
        cnpj: {}
        Endereço: {}{}, complemento: {}
        Email: {}
        Numero: {}'''.format(nomesoc,
            cnpj,
            local, num_endere_empre, comp_empre,
            email_empre,
            numero_tele_empre))

        #cria uma planilia com o excel com as informações
        try:
            workbook = load_workbook('BD_python.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(['Nome','cnpj','Endereço', 'Numero_do_endereço', 'Complemento','Email','Numero de telefone'])

        sheet.append([nomesoc, cnpj, local, num_endere_empre, comp_empre, email_empre,numero_tele_empre])

        workbook.save('BD_python.xlsx')
        break
    else:
        print("escolha invalida tente novamente")
        
        
    if esclog == "2":
        #solicita dados caso o usuario seja um cliente
        nome = input(str("digite seu nome: "))
            
        idade = input("digite sua idade: ")

        #cria uma lista de generos
        generos = ['Masculino', 'Feminino','Não quero declarar', 'outros']

        #cria um loop para certificar das escolhas do usuario
        while True:
            print('Selecione seu genero:')
            for i, genero in enumerate(generos, start=1):
                print(f"{i}. {genero}")
                
        #solicita a escolha do genero
            escolha = int(input("Digite o numero do genero selecionado: "))
            
        #verifica se a escolha é valida
            if 1 <= escolha <= len(generos):
                generos_selecionado = generos[escolha -1]
                print(f"Você selecionou {generos_selecionado}.")
                break
            else:
                print("escolha invalida. por favor, selecione um numero valido")

        #solicita o endereço
        endereço = input("Endereço: ")

        #solicita o numero do endereço
        numero_do_endereço =input("Numero: ")

        #cria uma lista para declara se tem ou não complemento
        lis_complemento = ['sim','não']

        #cria um loop para certificar das escolhas do usuario
        while True:
            print('Seu endereço tem complemento?')
            for i, complemento in enumerate(lis_complemento, start=1):
                print(f"{i}. {complemento}")
                
            #solicita uma escolha entre sim ou não
            escolha2 = int(input("digite 1 se tiver complemento, se não 2 "))
            if 1 <= escolha2 <= len(lis_complemento):
                complemento_selecionado = lis_complemento[escolha2 - 1]
                
                #caso o endereço precise de complemento 
                if complemento_selecionado == 'sim':
                    complemento = input("Digite seu complemento: ")
                    print(f"Endereço: {endereço}, {numero_do_endereço}, complemento: {complemento}")
                
                #caso o endereço não precise de um complemento
                else:
                    print(f"Endereço: {endereço}, {numero_do_endereço}")
                break
                    
            #caso o usuario for desprovido de inteligencia
            else:
                print("Escolha inválida. Por favor, selecione 1 ou 2")
        while True:
        #solicita um E-mail
            email = input("digite seu email: ")
            valid_email = re.compile(r'^[^\s]+@gmail.com$')
            if valid_email.match(email):
                print("email valido")
                break
            else:
                print("Formato do email invalido, Certifique-se de que não há espaços e que termina com @gmail.com.")
                
        #solicita um numéro de telefone
        numero_tele = input("Digite seu numero com ddd: ")
        
        #mensagem dizendo que finalizou o cadastro e refornecendo as informações
        print("Cadastro comcluido suas informações são:")
        print('''Nome: {}
        Idade: {}
        Endereço: {}{}, complemento: {}
        Email: {}
        Numero: {}'''.format(nome,
            idade,
            endereço, numero_do_endereço, complemento,
            email,
            numero_tele))

        #cria uma planilia com o excel com as informações
        try:
            workbook = load_workbook('BD_python.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(['Nome','Idade','Endereço', 'Numero_do_endereço', 'Complemento','Email','Numero de telefone'])

        sheet.append([nome, idade, endereço, numero_do_endereço, complemento, email ,numero_tele])

        workbook.save('BD_python.xlsx')
        break
    else:
        print("escolha invalida tente novamente")